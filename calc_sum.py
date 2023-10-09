import os
import tkinter.filedialog

import openpyxl
import datetime
from tkinter import *
import time
from openpyxl.styles import PatternFill

ASSET_ID = 4
ASSET_VALUE = 33
ASSET_CALC_VALUE = 66

maxlen = 15
delete_lib = [' ', '/', '(', ')', '（', '）', '^']


class Format():
    def __init__(self, init_window):
        self.asset_format_back_button = None
        self.asset_format_button = None
        self.start = None
        self.asset_calc_value = None
        self.asset_calc_value_title = None
        self.asset_value = None
        self.asset_value_title = None
        self.asset_id = None
        self.asset_id_title = None
        self.status = None
        self.output = None
        self.output_label = None
        self.input = None
        self.input_button = None
        self.input_label = None
        self.init_title = None
        self.init_window = init_window

    def set_init_window(self):
        self.init_window.title("设备总价值重新计算")
        self.init_window.geometry('700x300+10+10')
        # self.init_window.geometry('320x160+10+10')
        self.init_title = Label(self.init_window, text="Format Function")
        self.init_title.grid(row=0, column=0)
        self.input_label = Label(self.init_window, text="Input path")
        self.input_label.grid(row=1, column=0)
        self.input_button = Button(self.init_window, text="upload", command=self.upload_file)
        self.input_button.grid(row=1, column=11)
        self.input = Entry(self.init_window, width=66)
        self.input.grid(row=1, column=1, columnspan=10)
        self.output_label = Label(self.init_window, text="Output")
        self.output_label.grid(row=2, column=0)
        self.output = Entry(self.init_window, width=66)
        self.output.grid(row=2, column=1, columnspan=10)
        self.status = Label(self.init_window, text="Status")
        self.status.grid(row=17, column=0)
        self.status = Entry(self.init_window, width=7)
        self.status.grid(row=17, column=1)

        self.asset_id_title = Label(self.init_window, text="资产编码列数")
        self.asset_id_title.grid(row=4, column=1)
        self.asset_id = Entry(self.init_window, width=7)
        self.asset_id.grid(row=4, column=2)

        self.asset_value_title = Label(self.init_window, text="资产价值列数")
        self.asset_value_title.grid(row=4, column=3)
        self.asset_value = Entry(self.init_window, width=7)
        self.asset_value.grid(row=4, column=4)

        self.asset_calc_value_title = Label(self.init_window, text="计算后价值列数")
        self.asset_calc_value_title.grid(row=4, column=5)
        self.asset_calc_value = Entry(self.init_window, width=7)
        self.asset_calc_value.grid(row=4, column=6)

        self.start = Button(self.init_window, text="重新计算求和", bg="lightblue", width=10,
                            command=self.calc_value)  # 调用内部方法  加()为直接调用
        self.start.grid(row=13, columnspan=20)
        self.asset_format_button = Button(self.init_window, text="格式资产编码", bg="lightblue", width=10,
                                          command=self.format_asset_id)  # 调用内部方法  加()为直接调用
        self.asset_format_button.grid(row=12, columnspan=20)
        # self.asset_format_back_button = Button(self.init_window, text="反向编码", bg="lightblue", width=10,
        #                                           command=self.format_asset_id_back)
        # self.asset_format_back_button.grid(row=14, columnspan=20)
        print(self.input.get())

    def mergec(self, ws, begin, mid, end, col_id):
        i = begin
        j = mid + 1
        tmp = []
        while i <= mid and j <= end:
            if ws.cell(i, col_id).value < ws.cell(j, col_id).value:
                tmp.append(i)
                i += 1
            else:
                tmp.append(j)
                j += 1
        while i <= mid:
            tmp.append(i)
            i += 1
        while j <= end:
            tmp.append(j)
            j += 1
        tmp_wb = []
        for i in range(begin, end + 1):
            for j in range(1, ws.max_column + 1):
                tmp_wb.append(ws.cell(i, j).value)
        for i in range(begin, end + 1):
            for j in range(1, ws.max_column + 1):
                ws.cell(i, j).value = tmp_wb[(tmp[i - begin] - begin) * ws.max_column + j - 1]

    def merge_sort(self, ws, begin, end, col_id):
        if begin == end:
            return
        mid = (begin + end) // 2
        print(begin, mid, end)
        self.merge_sort(ws, begin, mid, col_id)
        self.merge_sort(ws, mid + 1, end, col_id)
        self.mergec(ws, begin, mid, end, col_id)

    def calc_value(self):
        self.status.delete(0, 'end')
        self.status.insert(0, "running")
        self.status['bg'] = 'yellow'
        inputfile = self.input.get()
        outputfile = self.output.get()
        asset_id = self.asset_id.get()
        asset_value = self.asset_value.get()
        asset_calc_value = self.asset_calc_value.get()
        if asset_id == '':
            asset_id = ASSET_ID
        if asset_value == '':
            asset_value = ASSET_VALUE
        if asset_calc_value == '':
            asset_calc_value = ASSET_CALC_VALUE
        asset_id = int(asset_id)
        asset_value = int(asset_value)
        asset_calc_value = int(asset_calc_value)
        if inputfile == '':
            print('请填写完整信息')
            self.status.delete(0, 'end')
            self.status.insert(0, "Error")
            self.status['bg'] = 'red'
        else:
            # 跳过第一行，从第二行开始读取
            # 以资产编码为排序的依据,从小到大进行排序
            # 读入数据
            wb = openpyxl.load_workbook(inputfile)
            ws = wb.active
            fill = PatternFill("solid", fgColor="FFFF00")
            # wb = wb[wb.sheetnames[0]]
            # print(wb[1][1].value)
            # get the max row number
            max_row = ws.max_row
            print("max_row", max_row)
            # print(max_row)
            # sort
            # for i in range(2, max_row + 1):
            #     print(i)
            #     for j in range(i + 1, max_row + 1):
            #         if int(ws.cell(row=i, column=int(asset_id)).value[:10]) > int(ws.cell(row=j, column=int(asset_id)).value[:10]):
            #             for k in range(1, ws.max_column + 1):
            #                 temp = ws.cell(row=i, column=k).value
            #                 ws.cell(row=i, column=k).value = ws.cell(row=j, column=k).value
            #                 ws.cell(row=j, column=k).value = temp
            #         if int(ws.cell(row=i, column=int(asset_id)).value[:10]) == int(ws.cell(row=j, column=int(asset_id)).value[:10]):
            #             if len(ws.cell(row=i, column=int(asset_id)).value) > len(ws.cell(row=j, column=int(asset_id)).value):
            #                 for k in range(1, ws.max_column + 1):
            #                     temp = ws.cell(row=i, column=k).value
            #                     ws.cell(row=i, column=k).value = ws.cell(row=j, column=k).value
            #                     ws.cell(row=j, column=k).value = temp
            #             if len(ws.cell(row=i, column=int(asset_id)).value) == len(ws.cell(row=j, column=int(asset_id)).value):
            #                 if ws.cell(row=i, column=int(asset_id)).value > ws.cell(row=j, column=int(asset_id)).value:
            #                     for k in range(1, ws.max_column + 1):
            #                         temp = ws.cell(row=i, column=k).value
            #                         ws.cell(row=i, column=k).value = ws.cell(row=j, column=k).value
            #                         ws.cell(row=j, column=k).value = temp
            # change to merge sort
            # self.merge_sort(ws, 2, max_row, int(asset_id))

            # calculate
            # 将资产编码前十位相同的资产价值相加，得到计算后价值
            i = 2
            while i < max_row:
                sum = 0.0
                # print(i)
                # print(ws.cell(row=44025, column=int(asset_id)).value[11])
                for j in range(i, max_row + 1):
                    # 如果第j行为空，跳出循环
                    if ws.cell(row=j, column=int(asset_id)).value is None:
                        break
                    if int(ws.cell(row=i, column=int(asset_id)).value[:10]) == int(
                            ws.cell(row=j, column=int(asset_id)).value[:10]):
                        sum = sum + float(ws.cell(row=j, column=int(asset_value)).value)
                        # if i != j:
                        #     print("calc", i, j, sum)
                    else:
                        break
                # print(i, ws.cell(row=i, column=int(asset_id)).value[11] == '0')
                if ws.cell(row=i, column=int(asset_id)).value is not None and not ws.cell(row=i, column=int(asset_id)).value[11] == '0':
                    # change the color into yellow
                    ws.cell(row=i, column=int(asset_id)).fill = fill

                ws.cell(row=i, column=int(asset_calc_value)).number_format = '0.00'
                ws.cell(row=i, column=int(asset_calc_value)).value = sum
                i = j
                if ws.cell(row=i, column=int(asset_id)).value is None or ws.cell(row=i, column=int(asset_id)).value == '':
                    break

            i = 2
            while i < max_row:
                if ws.cell(row=i, column=int(asset_id)).value is not None and ws.cell(row=i, column=int(asset_id)).value[11] == '0':
                    ws.cell(row=i, column=int(asset_id)).value = ws.cell(row=i, column=int(asset_id)).value[:10]
                i = i + 1

            if outputfile == '':
                wb.save("output.xlsx")
            else:
                wb.save(outputfile)

            self.status.delete(0, 'end')
            self.status.insert(0, "Success")
            self.status['bg'] = 'green'

            return

    def format_asset_id(self):
        # 如果是10位，后面加上'-0'
        # 如果是12位，继续
        self.status.delete(0, 'end')
        self.status.insert(0, "running")
        self.status['bg'] = 'yellow'
        inputfile = self.input.get()
        outputfile = self.output.get()
        asset_id = self.asset_id.get()
        asset_value = self.asset_value.get()
        asset_calc_value = self.asset_calc_value.get()
        if asset_id == '':
            asset_id = ASSET_ID
        if asset_value == '':
            asset_value = ASSET_VALUE
        if asset_calc_value == '':
            asset_calc_value = ASSET_CALC_VALUE
        asset_id = int(asset_id)
        asset_value = int(asset_value)
        asset_calc_value = int(asset_calc_value)
        if inputfile == '':
            print('请填写完整信息')
            self.status.delete(0, 'end')
            self.status.insert(0, "Error")
            self.status['bg'] = 'red'
        else:
            wb = openpyxl.load_workbook(inputfile)
            ws = wb.active
            max_row = ws.max_row
            i = 2
            while i < max_row:
                if ws.cell(row=i, column=int(asset_id)).value:
                    if len(str(ws.cell(row=i, column=int(asset_id)).value)) == 10:
                        ws.cell(row=i, column=int(asset_id)).value = str(ws.cell(row=i, column=int(asset_id)).value) + '-0'
                i = i + 1
            if outputfile == '':
                wb.save("output.xlsx")
            else:
                wb.save(outputfile)
            self.status.delete(0, 'end')
            self.status.insert(0, "Success")
            self.status['bg'] = 'green'
            return

    def format_asset_id_back(self):
        self.status.delete(0, 'end')
        self.status.insert(0, "running")
        self.status['bg'] = 'yellow'
        inputfile = self.input.get()
        outputfile = self.output.get()
        asset_id = self.asset_id.get()
        asset_value = self.asset_value.get()
        asset_calc_value = self.asset_calc_value.get()
        if asset_id == '':
            asset_id = ASSET_ID
        if asset_value == '':
            asset_value = ASSET_VALUE
        if asset_calc_value == '':
            asset_calc_value = ASSET_CALC_VALUE
        asset_id = int(asset_id)
        asset_value = int(asset_value)
        asset_calc_value = int(asset_calc_value)
        if inputfile == '':
            print('请填写完整信息')
            self.status.delete(0, 'end')
            self.status.insert(0, "Error")
            self.status['bg'] = 'red'
        else:
            wb = openpyxl.load_workbook(inputfile)
            ws = wb.active
            max_row = ws.max_row
            i = 2
            while i < max_row + 1:
                if ws.cell(row=i, column=int(asset_id)).value[11] == '0':
                    ws.cell(row=i, column=int(asset_id)).value = ws.cell(row=i, column=int(asset_id)).value[:10]
                i = i + 1
            if outputfile == '':
                wb.save("output.xlsx")
            else:
                wb.save(outputfile)
            self.status.delete(0, 'end')
            self.status.insert(0, "Success")
            self.status['bg'] = 'green'
            return

    def upload_file(self):
        self.input.delete(0, 'end')
        inputfile = tkinter.filedialog.askopenfilename()
        if inputfile[-4:] != 'xlsx':
            print('请上传excel表格')
        else:
            self.input.insert(0, inputfile)


def choose():
    # format_all()
    init_window = Tk()
    my_window = Format(init_window)
    my_window.set_init_window()
    init_window.mainloop()


choose()
